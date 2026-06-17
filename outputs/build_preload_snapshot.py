from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "fleet_apm_snapshot.js"

FILES = {
    "maintenance": Path(r"C:\Users\bahan\Downloads\Fleet_Preventative_Maintenance_&_Repair_Work_Orders_20260609.xlsx"),
    "inventory": Path(r"C:\Users\bahan\Downloads\Fleet_Inventory_20260611.xlsx"),
    "monthly": Path(r"C:\Users\bahan\Downloads\Fleet_Monthly_Costs_20260611.xlsx"),
}

FIELDS = {
    "maintenance_required": [
        "unique_work_order_no", "create_date", "work_order_yr", "job_type", "eq_equip_no",
        "work_order_status", "meter_1_reading", "datetime_out_service", "datetime_in_service",
        "datetime_open", "datetime_finished", "datetime_closed", "downtime_hrs_user",
        "downtime_hrs_shop", "reas_reas_for_repair", "reas_for_repair_desc",
        "dept_equip_dept", "dept_equip_dept_name", "meter_1_life_total",
        "labor_hours", "labor_cost", "parts_cost", "comml_cost", "total_cost",
    ],
    "maintenance_optional": [
        "warranty", "pri_priority_code", "loc_work_order_loc", "loc_work_order_loc_name",
        "datetime_unit_in", "datetime_due", "datetime_pm_sched",
    ],
    "inventory_required": [
        "eq_equip_no", "asset_type", "class_class_bench", "est_replace_cost",
        "manufacturer", "model", "description", "department_name", "dept_dept_code",
        "in_service_date", "last_meter_1_reading", "meter_1_at_delivery", "meter_1_type",
    ],
    "inventory_optional": [
        "delivery_date", "retire_date", "sale_date", "status_codes", "last_meter_1_date",
        "year", "class_class_maint", "class_class_pm", "class_class_shop_sch",
        "original_cost", "est_replace_yr", "est_replace_mo", "weight_gross",
        "wheelbase", "qty_axles", "cab_axle_len",
    ],
    "monthly_required": [
        "eq_equip_no", "cost_year", "cost_month", "meter_1_usage", "meter_1_eom",
        "fuel_cost", "cng_cost", "oil_cost", "misc_cost", "repair_labor_cost",
        "repair_parts_cost", "pm_labor_cost", "pm_parts_cost",
    ],
    "monthly_optional": [],
}


def norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def js_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def combined_fields(name: str) -> list[str]:
    return FIELDS[f"{name}_required"] + FIELDS.get(f"{name}_optional", [])


def read_selected(
    path: Path,
    required: list[str],
    optional: list[str] | None = None,
    row_filter=None,
) -> list[list[object]]:
    optional = optional or []
    wanted = required + optional
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [norm_header(v) for v in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    missing = [field for field in required if field not in index]
    if missing:
        raise RuntimeError(f"{path.name} missing expected columns: {missing}")

    selected = [(field, index.get(field)) for field in wanted]
    records: list[list[object]] = []
    for row in rows:
        item = []
        has_value = False
        for field, idx in selected:
            value = js_value(row[idx] if idx is not None and idx < len(row) else "")
            if value not in ("", None):
                has_value = True
            item.append(value)
        if has_value and (row_filter is None or row_filter(item, wanted)):
            records.append(item)
    wb.close()
    return records


def main() -> None:
    schema = {
        "maintenance": combined_fields("maintenance"),
        "inventory": combined_fields("inventory"),
        "monthly": combined_fields("monthly"),
    }

    maintenance = read_selected(
        FILES["maintenance"],
        FIELDS["maintenance_required"],
        FIELDS["maintenance_optional"],
    )
    asset_idx = schema["maintenance"].index("eq_equip_no")
    work_order_assets = {str(row[asset_idx]) for row in maintenance if row[asset_idx] not in ("", None)}

    monthly_asset_idx = schema["monthly"].index("eq_equip_no")
    monthly_cost_idxs = [
        schema["monthly"].index(field)
        for field in [
            "fuel_cost",
            "cng_cost",
            "oil_cost",
            "misc_cost",
            "repair_labor_cost",
            "repair_parts_cost",
            "pm_labor_cost",
            "pm_parts_cost",
        ]
    ]

    def keep_monthly(row: list[object], fields: list[str]) -> bool:
        if str(row[monthly_asset_idx]) not in work_order_assets:
            return False
        return any(float(row[idx] or 0) != 0 for idx in monthly_cost_idxs)

    snapshot = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": "local XLSX preload",
        "schema": schema,
        "maintenance": maintenance,
        "inventory": read_selected(
            FILES["inventory"],
            FIELDS["inventory_required"],
            FIELDS["inventory_optional"],
        ),
        "monthly": read_selected(
            FILES["monthly"],
            FIELDS["monthly_required"],
            FIELDS["monthly_optional"],
            keep_monthly,
        ),
    }
    text = "window.FLEET_APM_PRELOAD = "
    text += json.dumps(snapshot, ensure_ascii=False, separators=(",", ":"))
    text += ";\n"
    OUTPUT.write_text(text, encoding="utf-8")
    print(json.dumps({k: len(v) for k, v in snapshot.items() if isinstance(v, list)}, indent=2))
    print(OUTPUT)


if __name__ == "__main__":
    main()
