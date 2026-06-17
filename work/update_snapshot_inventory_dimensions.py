from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SNAPSHOT = ROOT / "outputs" / "fleet_apm_snapshot.js"
INVENTORY_XLSX = Path(r"C:\Users\bahan\Downloads\Fleet_Inventory_20260611.xlsx")
DIM_FIELDS = ["weight_gross", "wheelbase", "qty_axles", "cab_axle_len"]
PREFIX = "window.FLEET_APM_PRELOAD = "


def norm_header(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def js_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def load_inventory_dimensions() -> dict[str, dict[str, object]]:
    wb = load_workbook(INVENTORY_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [norm_header(v) for v in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    asset_idx = index["eq_equip_no"]
    dim_index = {field: index.get(field) for field in DIM_FIELDS}
    values: dict[str, dict[str, object]] = {}
    for row in rows:
        asset = str(row[asset_idx] or "").strip()
        if not asset:
            continue
        values[asset] = {
            field: js_value(row[idx] if idx is not None and idx < len(row) else "")
            for field, idx in dim_index.items()
        }
    wb.close()
    return values


def main() -> None:
    text = SNAPSHOT.read_text(encoding="utf-8")
    if not text.startswith(PREFIX):
        raise RuntimeError("Unexpected snapshot format")
    payload = text[len(PREFIX):].strip()
    if payload.endswith(";"):
        payload = payload[:-1]
    snapshot = json.loads(payload)
    fields = snapshot["schema"]["inventory"]
    asset_idx = fields.index("eq_equip_no")
    dim_values = load_inventory_dimensions()

    for field in DIM_FIELDS:
        if field not in fields:
            fields.append(field)
            for row in snapshot["inventory"]:
                asset = str(row[asset_idx] or "").strip()
                row.append(dim_values.get(asset, {}).get(field, ""))

    snapshot["generated_at"] = datetime.now().isoformat(timespec="seconds")
    SNAPSHOT.write_text(PREFIX + json.dumps(snapshot, ensure_ascii=False, separators=(",", ":")) + ";\n", encoding="utf-8")
    print(json.dumps({
        "inventory_fields": len(snapshot["schema"]["inventory"]),
        "added_fields": [field for field in DIM_FIELDS if field in snapshot["schema"]["inventory"]],
        "inventory_rows": len(snapshot["inventory"]),
    }, indent=2))


if __name__ == "__main__":
    main()
